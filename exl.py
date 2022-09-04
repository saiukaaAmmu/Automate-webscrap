import openpyxl as o
from openpyxl import Workbook, load_workbook


class exl():
    #def __init__(self):

    def user_create_workbook(self):
        #self.wb=Workbook()
        try:
            self.wb = load_workbook("reg.xlsx")
        except:
            self.wb = Workbook()
        self.sheet=self.wb.active
        self.sheet["A1"]=" User Name "
        self.sheet["B1"]=" Password "
        self.sheet["C1"]=" email id "

    def add_data(self,username:str,password:str, mail:str):
        __col_num = str(self.sheet.max_row+1)
        self.sheet["A"+__col_num]=username
        self.sheet["B"+__col_num]=password
        self.sheet["C"+__col_num]=mail

        self.wb.save('reg.xlsx')
        return True
